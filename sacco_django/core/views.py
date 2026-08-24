import bcrypt
import hashlib
import io
import json
import requests
import csv
import qrcode
import base64
from io import BytesIO
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from datetime import datetime
from django.db import connection, transaction
from django.db.models import Sum, Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views.decorators.http import require_POST
from django.http import FileResponse, HttpResponse
from django.template.loader import render_to_string
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from .models import Membres, Groupes as Groupe, Partenaire, Pret, TransactionHistory, Pret, DemandeCredit, HistoriqueEpargne
from .forms import TransactionForm, EmployeeCreationForm, LoanRequestForm
from reportlab.pdfgen import canvas
from django.contrib.auth.decorators import permission_required, user_passes_test, login_required
from django.utils import timezone
from .utils import calculer_amortissement
from django.conf import settings
from django.contrib.auth import authenticate, login
from django.contrib.auth.forms import AuthenticationForm
from django.views.decorators.cache import never_cache
from django.views.decorators.csrf import ensure_csrf_cookie
from django.contrib.staticfiles import finders
from bs4 import BeautifulSoup
from passlib.context import CryptContext
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from xhtml2pdf import pisa
from .decorators import partner_required
from django.utils.http import url_has_allowed_host_and_scheme


try:
    from xhtml2pdf import pisa
except ImportError:
    pisa = None
try:
    import openpyxl
except ImportError:
    openpyxl = None


def is_manager(user):
    return user.is_superuser or user.groups.filter(name='Managers').exists()
def is_agent_credit(user):
    return user.is_superuser or user.groups.filter(name='AgentsCredit').exists()
def is_partner(user):
    return user.is_superuser or user.groups.filter(name='Partenaires').exists()


def redirect_based_on_role(user):
    if user.is_superuser or user.is_staff or user.groups.filter(name='Managers').exists():
        return redirect('manager_dashboard')
    elif user.groups.filter(name='AgentsCredit').exists():
        return redirect('agent_dashboard')
    elif user.groups.filter(name='Partenaires').exists():
        return redirect('partner_dashboard')
    return redirect('dashboard')


@ensure_csrf_cookie
@never_cache
def login_view(request):
    next_url = request.POST.get('next') or request.GET.get('next')

    def get_redirect_url(default_route):
        """Valide l'URL 'next' ou renvoie la route par défaut."""
        if next_url and url_has_allowed_host_and_scheme(
            url=next_url,
            allowed_hosts={request.get_host()},
            require_https=request.is_secure()
        ):
            return next_url
        return redirect(default_route).url

    if request.user.is_authenticated:
        if request.user.groups.filter(name='Partenaires').exists():
            return redirect(get_redirect_url('core:partner_dashboard'))
        elif request.user.is_superuser or request.user.is_staff:
            return redirect(get_redirect_url('core:manager_dashboard'))
        else:
            logout(request)

    if 'membre_id' in request.session:
        return redirect(get_redirect_url('dashboard'))

    if request.method == 'POST':
        user_type = request.POST.get('user_type')
        identifier = request.POST.get('identifier', '').strip()
        secret = request.POST.get('secret', '').strip()

        if not identifier and request.POST.get('telephone'):
            identifier = request.POST.get('telephone')
            user_type = 'membre'
        if not secret and request.POST.get('pin'):
            secret = request.POST.get('pin')

        if user_type == 'membre':
            try:
                membre = Membres.objects.get(telephone=identifier)
                stored_pin = str(membre.pin)
                authenticated = False

                if stored_pin.startswith('$2b$') or stored_pin.startswith('$2a$'):
                    try:
                        if bcrypt.checkpw(secret.encode('utf-8'), stored_pin.encode('utf-8')):
                            authenticated = True
                    except Exception:
                        pass
                elif len(stored_pin) == 64:
                    if hashlib.sha256(secret.encode('utf-8')).hexdigest() == stored_pin:
                        authenticated = True
                else:
                    if stored_pin == secret:
                        authenticated = True

                if authenticated:
                    membre.last_login = timezone.now().strftime("%Y-%m-%d %H:%M:%S")
                    membre.save()
                    request.session['user_id'] = membre.id
                    request.session['user_type'] = 'membre'
                    request.session['membre_id'] = membre.id
                    request.session['membre_nom'] = f"{membre.prenom} {membre.nom}"
                    request.session['role'] = membre.role
                    return redirect(get_redirect_url('core:dashboard'))
                else:
                    return render(request, 'core/login.html', {'error_message': 'Téléphone ou Code PIN incorrect.', 'next': next_url})
            except Membres.DoesNotExist:
                return render(request, 'core/login.html', {'error_message': 'Téléphone ou Code PIN incorrect.', 'next': next_url})
        else:
            user = authenticate(request, username=identifier, password=secret)
            if user is not None and user.is_active:
                is_partner = user.groups.filter(name='Partenaires').exists()
                is_admin = user.is_superuser or user.is_staff

                if user_type == 'partenaire' and is_partner:
                    login(request, user)
                    request.session['user_id'] = user.id
                    request.session['user_type'] = 'partenaire'
                    request.session['role'] = 'partenaire'
                    return redirect(get_redirect_url('core:partner_dashboard'))
                elif user_type == 'admin' and is_admin:
                    login(request, user)
                    request.session['user_id'] = user.id
                    request.session['user_type'] = 'admin'
                    request.session['role'] = 'admin'
                    return redirect(get_redirect_url('core:manager_dashboard'))
                else:
                    return render(request, 'core/login.html',
                                  {'error_message': "Vous n'avez pas les droits pour cet espace.", 'next': next_url})
            else:
                return render(request, 'core/login.html',
                              {'error_message': 'Nom d\'utilisateur ou mot de passe incorrect.', 'next': next_url})

    return render(request, 'core/login.html', {'next': next_url})

universal_login_view = login_view


def partner_login_view(request):
    error_message = None
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)
        if user is not None and user.is_active:
            if user.groups.filter(name='Partenaires').exists() or user.is_superuser:
                login(request, user)
                return redirect('core:partner_dashboard')
            else:
                error_message = "Vous n'avez pas les droits d'accès partenaires."
        else:
            error_message = "Nom d'utilisateur ou mot de passe incorrect."

    return render(request, 'core/partner_login.html', {'error_message': error_message})


def manager_dashboard_view(request):
    is_admin_auth = request.user.is_authenticated and (
            request.user.is_superuser or request.user.is_staff or is_manager(request.user)
    )
    role = str(request.session.get('role', '')).lower()
    is_session_admin = 'membre_id' in request.session and ('admin' in role or 'gestionnaire' in role)
    if not is_admin_auth and not is_session_admin:
        return redirect('login')

    if request.method == 'POST' and 'update_role' in request.POST:
        membre_id = request.POST.get('membre_id')
        nouveau_role = request.POST.get('nouveau_role')
        try:
            membre_a_modifier = Membres.objects.get(id=membre_id)
            membre_a_modifier.role = nouveau_role
            membre_a_modifier.save()
            messages.success(request, f"Le rôle de {membre_a_modifier.nom} a été mis à jour avec succès.")
        except Membres.DoesNotExist:
            messages.error(request, "Membre introuvable.")
        return redirect(f"{request.path}?gid={request.GET.get('gid', 1)}")

    selected_gid = request.GET.get('selected_gid', request.GET.get('gid', ''))
    selected_member_id = request.GET.get('selected_member_id', '')
    mois_filtre = request.GET.get('mois_filtre', datetime.now().strftime('%Y-%m'))

    try:
        annee, mois = map(int, mois_filtre.split('-'))
    except ValueError:
        mois_filtre = datetime.now().strftime('%Y-%m')
        annee, mois = map(int, mois_filtre.split('-'))

    prefixe_mois = f"{annee}-{mois:02d}"
    membres_qs = Membres.objects.filter(is_active=True)
    if selected_gid:
        membres_qs = membres_qs.filter(groupe_id=selected_gid)
    if selected_member_id:
        membres_qs = membres_qs.filter(id=selected_member_id)

    query = request.GET.get('q', '').strip()
    if query:
        tous_les_membres = Membres.objects.filter(
            Q(nom__icontains=query) |
            Q(prenom__icontains=query) |
            Q(telephone__icontains=query)
        ).order_by('nom')
    else:
        tous_les_membres = Membres.objects.all().order_by('nom')

    role_query = request.GET.get('role_q', '').strip()
    if role_query:
        membres_pour_role = Membres.objects.filter(
            Q(id__icontains=role_query) |
            Q(nom__icontains=role_query) |
            Q(prenom__icontains=role_query) |
            Q(telephone__icontains=role_query)
        ).order_by('nom')[:15]
    else:
        membres_pour_role = Membres.objects.all().order_by('nom')[:15]

    page_number = request.GET.get('page', 1)
    paginator = Paginator(tous_les_membres, 10)
    try:
        membres_pagines = paginator.page(page_number)
    except PageNotAnInteger:
        membres_pagines = paginator.page(1)
    except EmptyPage:
        membres_pagines = paginator.page(paginator.num_pages)

    dates_reunions_qs = HistoriqueEpargne.objects.filter(
        date_reunion__startswith=prefixe_mois
    )
    if selected_gid:
        dates_reunions_qs = dates_reunions_qs.filter(groupe_id=selected_gid)

    dates_reunions = list(
        dates_reunions_qs.values_list('date_reunion', flat=True).distinct().order_by('date_reunion')
    )

    pivot_sociale = []
    pivot_presences = []
    table_epargne = []
    total_report_anterieur = 0
    total_cotisations_mois = 0
    total_cumul_general = 0
    totaux_par_colonne_dict = {date: 0 for date in dates_reunions}
    total_general_caisse_sociale = 0
    grand_total_general = 0

    for m in membres_qs:
        report_anterieur = HistoriqueEpargne.objects.filter(
            membre_id=m.id,
            date_reunion__lt=f"{prefixe_mois}-01"
        ).aggregate(Sum('caisse_sociale'))['caisse_sociale__sum'] or 0

        montants_mois = []
        statuts_list = []
        details_epargne = []
        sum_mois = 0
        cumul_epargne = 0

        total_caisse_sociale_membre = HistoriqueEpargne.objects.filter(
            membre_id=m.id,
            date_reunion__startswith=prefixe_mois
        ).aggregate(Sum('caisse_sociale'))['caisse_sociale__sum'] or 0

        for d in dates_reunions:
            he = HistoriqueEpargne.objects.filter(membre_id=m.id, date_reunion=d).first()
            valeur_sociale = he.caisse_sociale if (he and he.caisse_sociale) else 0
            valeur_epargne = he.epargne if (he and he.epargne) else 0
            statut = he.status_presence if he else '-'
            montants_mois.append(valeur_sociale)
            sum_mois += valeur_sociale
            statuts_list.append(statut)
            cumul_epargne += valeur_epargne
            details_epargne.append(f"{valeur_epargne:,.0f} / {cumul_epargne:,.0f}")
            totaux_par_colonne_dict[d] += valeur_epargne

        cumul_actuel = report_anterieur + sum_mois
        pivot_sociale.append({
            'id': m.pk, 'nom_complet': f"{m.prenom or ''} {m.nom}", 'report_anterieur': report_anterieur,
            'montants': montants_mois, 'total_mois': sum_mois, 'cumul_actuel': cumul_actuel,
        })

        pivot_presences.append({
            'id': m.id, 'nom_complet': f"{m.prenom or ''} {m.nom}", 'statuts': statuts_list
        })

        solde_total_avec_sociale = cumul_epargne + total_caisse_sociale_membre

        table_epargne.append({
            'id': m.id, 'nom_complet': f"{m.prenom or ''} {m.nom}", 'details': details_epargne,
            'caisse_sociale': total_caisse_sociale_membre,
            'solde_total_avec_sociale': f"{solde_total_avec_sociale:,.0f}",
        })

        total_report_anterieur += report_anterieur
        total_cotisations_mois += sum_mois
        total_cumul_general += cumul_actuel
        total_general_caisse_sociale += total_caisse_sociale_membre
        grand_total_general += solde_total_avec_sociale

    totaux_par_colonne = [f"{totaux_par_colonne_dict[d]:,.0f}" for d in dates_reunions]
    total_epargne = sum(m.solde_epargne or 0 for m in Membres.objects.all())
    total_caisse_sociale = sum(m.caisse_sociale or 0 for m in Membres.objects.all())
    total_credits_en_cours = sum(m.credit_en_cours or 0 for m in Membres.objects.all())
    top_membres = Membres.objects.all().order_by('-solde_epargne')[:10]
    noms_membres = [f"{m.prenom} {m.nom}" for m in top_membres]
    soldes_epargne = [float(m.solde_epargne or 0) for m in top_membres]
    brb_rates = get_brb_exchange_rates()
    tous_les_groupes = Groupe.objects.filter(est_archive=False)
    groupes_disponibles = tous_les_groupes
    historique_raw = HistoriqueEpargne.objects.filter(date_reunion__startswith=prefixe_mois)
    if selected_gid:
        historique_raw = historique_raw.filter(groupe_id=selected_gid)

    total_presences_effectives = 0
    total_absences_effectives = 0
    for h in historique_raw:
        statut_val = str(h.status_presence).strip().lower()
        if statut_val in ['p', 'présent', 'present', '1', 'oui']:
            total_presences_effectives += 1
        elif statut_val in ['a', 'absent', '0', 'non']:
            total_absences_effectives += 1

    total_appels = total_presences_effectives + total_absences_effectives
    taux_presence_global = round((total_presences_effectives / total_appels * 100), 1) if total_appels > 0 else 0

    formatted_dates = [
        d.strftime('%d/%m') if hasattr(d, 'strftime') else str(d)[-2:] + '/' + str(d)[5:7]
        for d in dates_reunions
    ]

    context = {
        'admin_nom': request.session.get('membre_nom') or request.user.username,
        'membres': membres_pagines,
        'total_membres': Membres.objects.count(),
        'total_epargne': total_epargne,
        'total_caisse_sociale': total_caisse_sociale,
        'total_credits': total_credits_en_cours,
        'search_query': query,
        'noms_membres': noms_membres,
        'soldes_epargne': soldes_epargne,
        'brb_rates': brb_rates,
        'total_groupes_actifs': tous_les_groupes.count(),
        'selected_gid': selected_gid,
        'selected_member_id': selected_member_id,
        'mois_filtre_courant': mois_filtre,
        'dates_reunions': formatted_dates,
        'pivot_presences': pivot_presences,
        'pivot_sociale': pivot_sociale,
        'table_epargne': table_epargne,
        'totaux_par_colonne': totaux_par_colonne,
        'total_general_caisse_sociale': f"{total_general_caisse_sociale:,.0f}",
        'grand_total_general': f"{grand_total_general:,.0f}",
        'total_report_anterieur': total_report_anterieur,
        'total_cotisations_mois': total_cotisations_mois,
        'total_cumul_general': total_cumul_general,
        'membres_pour_role': membres_pour_role,
        'role_query': role_query,
        'total_presences_effectives': total_presences_effectives,
        'total_absences_effectives': total_absences_effectives,
        'total_appels': total_appels,
        'taux_presence_global': taux_presence_global,
        'membres_groupe': tous_les_membres,
        'tous_les_membres': Membres.objects.filter(is_active=True),
        'tous_les_groupes': tous_les_groupes,
        'groupes_disponibles': groupes_disponibles,  # Ajouté ici
        'tous_les_partenaires': Partenaire.objects.all(),
    }

    return render(request, 'core/manager_dashboard.html', context)


def logout_view(request):
    request.session.flush()
    return redirect('/')


def member_profile_view(request):
    membre_id = request.session.get('membre_id') or request.session.get('user_id')
    if not membre_id:
        return redirect('login')

    membre = get_object_or_404(Membres, id=membre_id)
    context = {
        'membre': membre, 'admin_nom': request.session.get('membre_nom'),
    }
    return render(request, 'core/member_profile.html', context)


def member_qr_view(request):
    """Génère un QR Code unique pour identifier le membre."""
    membre_id = request.session.get('membre_id')
    if not membre_id:
        return redirect('login')

    membre = get_object_or_404(Membres, id=membre_id)
    qr_data = f"SACCO-ID:{membre.id} | Nom: {membre.nom} {membre.prenom} | Tél: {membre.telephone}"
    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(qr_data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="rgb(37, 99, 235)", back_color="white")  # Design bleu assorti
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    qr_image_base64 = base64.b64encode(buffer.getvalue()).decode()

    return render(request, 'core/member_qr.html', {'membre': membre, 'qr_image': qr_image_base64})


def ai_predictions_view(request):
    membre_id = request.session.get('membre_id')
    if not membre_id:
        return redirect('login')

    membre = get_object_or_404(Membres, id=membre_id)
    epargne_actuelle = membre.solde_epargne or 0
    epargne_predite_6m = epargne_actuelle * (1.05 ** 6)
    montant_simule = request.GET.get('montant_simule')
    duree_simulee = request.GET.get('duree_simulee')

    mensualite_estimee = 0
    evaluation_simulation = "Faible"

    if montant_simule and duree_simulee:
        try:
            montant_val = float(montant_simule)
            duree_val = int(duree_simulee)
            if duree_val > 0:
                total_avec_interet = montant_val * 1.05
                mensualite_estimee = round(total_avec_interet / duree_val, 2)

                if montant_val > (epargne_actuelle * 3):
                    evaluation_simulation = "Élevé (Dépasse 3x votre épargne)"
                else:
                    evaluation_simulation = "Faible (Conforme à votre profil)"
        except ValueError:
            pass

    conseil_ia = "Votre dynamique d'épargne est stable. Utilisez le simulateur ci-dessus pour tester l'impact d'un nouvel emprunt sur vos finances."

    context = {
        'membre': membre,
        'epargne_predite_6m': round(epargne_predite_6m, 2),
        'niveau_risque': "Faible",
        'score_fiabilite': 95,
        'conseil_ia': conseil_ia,
        'montant_simule': montant_simule or '',
        'duree_simulee': duree_simulee or '',
        'mensualite_estimee': mensualite_estimee,
        'evaluation_simulation': evaluation_simulation,
    }
    return render(request, 'core/ai_predictions.html', context)


def add_transaction_view(request, membre_id=None):
    role = str(request.session.get('role', '')).lower()
    is_staff_admin = request.user.is_authenticated and (request.user.is_superuser or request.user.is_staff)

    if not is_staff_admin and (
            'membre_id' not in request.session or ('admin' not in role and 'gestionnaire' not in role)):
        return redirect('login')

    membre = None
    if membre_id:
        membre = get_object_or_404(Membres, id=membre_id)

    if request.method == 'POST':
        form = TransactionForm(request.POST)
        if form.is_valid():
            transaction = form.save(commit=False)
            if membre and not transaction.membre:
                transaction.membre = membre
            transaction.save()
            return redirect('financial_report')
    else:
        initial_data = {'membre': membre} if membre else {}
        form = TransactionForm(initial=initial_data)

    return render(
        request,
        'core/add_transaction.html',
        {'form': form, 'membre': membre},
    )


def add_member_view(request):
    role = str(request.session.get('role', '')).lower()
    if 'membre_id' not in request.session or ('admin' not in role and 'gestionnaire' not in role):
        return redirect('login')
    error = None
    if request.method == 'POST':
        nom = request.POST.get('nom', '').strip()
        prenom = request.POST.get('prenom', '').strip()
        telephone = request.POST.get('telephone', '').strip()
        pin = request.POST.get('pin', '').strip()
        role_membre = request.POST.get('role', 'membre').strip()
        if not nom or not telephone or not pin:
            error = "Veuillez remplir tous les champs obligatoires."
        elif Membres.objects.filter(telephone=telephone).exists():
            error = "Ce numéro de téléphone est déjà utilisé par un autre membre."
        else:
            Membres.objects.create(nom=nom, prenom=prenom, telephone=telephone, pin=pin, role=role_membre, solde_epargne=0.0,
                caisse_sociale=0.0, credit_en_cours=0.0)
            return redirect('manager_dashboard')
    return render(request, 'core/add_member.html', {'error': error})


def grant_credit_view(request, membre_id):
    role = str(request.session.get('role', '')).lower()
    is_staff_admin = request.user.is_authenticated and (request.user.is_superuser or request.user.is_staff)

    if not is_staff_admin and (
            'membre_id' not in request.session or ('admin' not in role and 'gestionnaire' not in role)):
        return redirect('login')

    try:
        membre = Membres.objects.get(id=membre_id)
    except Membres.DoesNotExist:
        return redirect('manager_dashboard')

    if request.method == 'POST':
        try:
            montant = float(request.POST.get('montant_credit', 0))
            if montant > 0:
                membre.credit_en_cours = float(membre.credit_en_cours or 0) + montant
                membre.save()
                TransactionHistory.objects.create(
                    membre=membre, type_operation="Octroi de Crédit", montant=montant
                )
                return redirect('member_detail', membre_id=membre.id)
        except ValueError:
            pass

    return render(request, 'core/add_credit.html', {'membre': membre})


def member_detail_view(request, membre_id):
    role = str(request.session.get('role', '')).lower()
    if 'membre_id' not in request.session or ('admin' not in role and 'gestionnaire' not in role):
        return redirect('login')
    try:
        membre = Membres.objects.get(id=membre_id)
        transactions = membre.transactions.all().order_by('-date_transaction')
    except Membres.DoesNotExist:
        return redirect('manager_dashboard')
    context = {
        'membre': membre, 'transactions': transactions,
    }
    return render(request, 'core/member_detail.html', context)


def export_members_pdf(request):
    role = str(request.session.get('role', '')).lower()
    if 'membre_id' not in request.session or ('admin' not in role and 'gestionnaire' not in role):
        return redirect('login')

    search_query = request.GET.get('q', '')
    membres_qs = Membres.objects.all().order_by('nom')

    if search_query:
        membres_qs = membres_qs.filter(
            Q(nom__icontains=search_query) |
            Q(prenom__icontains=search_query) |
            Q(telephone__icontains=search_query)
        )

    context = {'membres': membres_qs}
    html_string = render_to_string('pdf/membres_pdf.html', context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="resume_membres_sacco.pdf"'
    pisa_status = pisa.CreatePDF(
        html_string, dest=response, link_callback=link_callback
    )

    if pisa_status.err:
        return HttpResponse('Erreur lors de la génération du PDF', status=500)

    return response


def loan_calculator_view(request):
    role = str(request.session.get('role', '')).lower()
    if 'membre_id' not in request.session or ('admin' not in role and 'gestionnaire' not in role):
        return redirect('login')
    schedule = []
    montant = 0
    taux_saisi = ""
    erreur = ""
    if request.method == 'POST':
        try:
            montant = float(request.POST.get('montant', 0))
            taux_mensuel = float(request.POST.get('taux', 0)) / 100.0
            mois = 3
            if montant > 0:
                if taux_mensuel > 0:
                    mensualite = (montant * taux_mensuel) / (1 - (1 + taux_mensuel) ** -mois)
                    solde_restant = montant
                    for i in range(1, mois + 1):
                        interet = solde_restant * taux_mensuel
                        capital_rembourse = mensualite - interet
                        solde_restant -= capital_rembourse
                        schedule.append({
                            'mois': i, 'mensualite': round(mensualite, 2), 'interet': round(interet, 2),
                            'capital': round(capital_rembourse, 2), 'solde': round(max(0, solde_restant), 2)
                        })
                else:
                    mensualite = montant / mois
                    for i in range(1, mois + 1):
                        schedule.append({
                            'mois': i, 'mensualite': round(mensualite, 2), 'interet': 0, 'capital': round(mensualite, 2),
                            'solde': round(montant - (mensualite * i), 2)
                        })
            taux_saisi = request.POST.get('taux', '')
        except ValueError:
            erreur = "Veuillez entrer des valeurs numériques valides."
    context = {
        'admin_nom': request.session.get('membre_nom'), 'schedule': schedule, 'montant_saisi': montant if montant > 0 else '',
        'taux_saisi': taux_saisi, 'erreur': erreur}
    return render(request, 'core/loan_calculator.html', context)

@login_required(login_url='/partenaire/login/')
@user_passes_test(is_partner, login_url='/partenaire/login/')
def partner_dashboard_view(request):
    context = {
        'partner_name': request.user.username,
    }
    return render(request, 'core/partner_dashboard.html', context)

@login_required
def partner_members_view(request):
    context = {
        'partner_name': request.user.username,
    }
    return render(request, 'core/partner_members.html', context)

@login_required
def partner_reports_view(request):
    context = {
        'partner_name': request.user.username,
    }
    return render(request, 'core/partner_reports.html', context)

@user_passes_test(is_manager, login_url='/login/')
def list_pending_loans_view(request):
    prets_en_attente = Pret.objects.filter(statut='EN_ATTENTE').select_related('membre')
    return render(request, 'core/pending_loans.html', {'prets': prets_en_attente})

@user_passes_test(is_manager, login_url='/login/')
def update_loan_status_view(request, pret_id, action):
    pret = get_object_or_404(Pret, id=pret_id)
    if action == 'approuver':
        pret.statut = 'APPROUVE'
        pret.date_approbation = timezone.now()
        messages.success(request, f"Le prêt #{pret.id} de {pret.membre.nom} a été approuvé avec succès.")
    elif action == 'rejeter':
        pret.statut = 'REJETE'
        pret.date_approbation = timezone.now()
        messages.warning(request, f"Le prêt #{pret.id} de {pret.membre.nom} a été rejeté.")
    pret.save()
    return redirect('pending_loans')


def generate_loan_contract_pdf(request, transaction_id):
    role = str(request.session.get('role', '')).lower()
    if 'membre_id' not in request.session or ('admin' not in role and 'gestionnaire' not in role):
        return redirect('login')
    transaction = get_object_or_404(TransactionHistory, id=transaction_id, type_operation="Octroi de Crédit")
    membre = transaction.membre
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle', parent=styles['Heading1'], fontSize=16,
        textColor=colors.HexColor('#2c3e50'), spaceAfter=20, alignment=1
    )
    body_style = ParagraphStyle(
        'BodyStyle', parent=styles['Normal'], fontSize=11,
        leading=16, spaceAfter=15
    )
    elements.append(Paragraph("CONTRAT DE PRÊT - SACCO FINTECH", title_style))
    elements.append(Spacer(1, 10))
    date_formatee = transaction.date_transaction.strftime('%d/%m/%Y')
    texte_contrat = f"""
    <b>Entre les soussignés :</b><br/><br/>
    D'une part, <b>SACCO Connect</b> (ci-après dénommé "le Prêteur"),<br/>
    Et d'autre part, <b>{membre.nom} {membre.prenom}</b>, titulaire du numéro de téléphone {membre.telephone} (ci-après dénommé "l'Emprunteur").<br/><br/>
    <b>Article 1 : Objet du contrat</b><br/>
    Le Prêteur consent à l'Emprunteur, qui accepte, un prêt d'un montant total de <b>{transaction.montant} BIF</b>.<br/><br/>
    <b>Article 2 : Modalités de remboursement</b><br/>
    Ce prêt, enregistré dans le système le {date_formatee}, est soumis aux conditions de remboursement et taux d'intérêt convenus dans le règlement intérieur de la coopérative. L'Emprunteur s'engage à rembourser l'intégralité du capital ainsi que les intérêts selon l'échéancier établi.<br/><br/>
    <b>Article 3 : Statut du prêt</b><br/>
    Statut actuel enregistré : <b>{transaction.get_statut_display()}</b>.<br/><br/>
    <b>Article 4 : Engagements</b><br/>
    En cas de non-remboursement aux dates convenues, l'Emprunteur s'expose aux pénalités de retard prévues par les statuts de la coopérative.
    """
    elements.append(Paragraph(texte_contrat, body_style))
    elements.append(Spacer(1, 60))
    signature_text = """
    <b>Signature du Prêteur</b> &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; <b>Signature de l'Emprunteur</b><br/>
    <font size="9">(Précédée de la mention "Lu et approuvé")</font>
    """
    elements.append(Paragraph(signature_text, body_style))
    doc.build(elements)
    buffer.seek(0)
    filename = f'Contrat_Pret_{membre.nom}_{transaction.id}.pdf'
    return FileResponse(buffer, as_attachment=True, filename=filename)


def apply_penalty_view(request, membre_id):
    role = str(request.session.get('role', '')).lower()
    if 'membre_id' not in request.session or ('admin' not in role and 'gestionnaire' not in role):
        return redirect('login')
    membre = get_object_or_404(Membres, id=membre_id)
    if request.method == 'POST':
        try:
            montant_penalite = float(request.POST.get('montant_penalite', 0))
            if montant_penalite > 0:
                TransactionHistory.objects.create(
                    membre=membre, type_operation='Penalite Retard', montant=montant_penalite, statut='ACTIF'
                )
                membre.credit_en_cours = float(membre.credit_en_cours or 0) + montant_penalite
                membre.save()
                messages.success(request, f"Pénalité de {montant_penalite} BIF appliquée avec succès à {membre.nom}.")
            else:
                messages.error(request, "Le montant de la pénalité doit être supérieur à zéro.")
        except ValueError:
            messages.error(request, "Valeur invalide.")
    return redirect('member_detail', membre_id=membre.id)


def financial_report_view(request):
    role = str(request.session.get('role', '')).lower()
    if 'membre_id' not in request.session or ('admin' not in role and 'gestionnaire' not in role):
        return redirect('login')

    transactions = TransactionHistory.objects.all().select_related('membre').order_by('-date_operation')
    type_op = request.GET.get('type_op', '').strip()
    if type_op:
        transactions = transactions.filter(type_operation=type_op)

    query = request.GET.get('q', '').strip()
    if query:
        transactions = transactions.filter(
            Q(membre__nom__icontains=query) |
            Q(membre__prenom__icontains=query) |
            Q(type_operation__icontains=query) |
            Q(effectue_par__icontains=query)
        )

    total_depots = TransactionHistory.objects.filter(type_operation='DEPOT').aggregate(Sum('montant'))['montant__sum'] or 0
    total_retraits = TransactionHistory.objects.filter(type_operation='RETRAIT').aggregate(Sum('montant'))['montant__sum'] or 0
    total_penalites = TransactionHistory.objects.filter(type_operation='PENALITE').aggregate(Sum('montant'))['montant__sum'] or 0
    total_credits = TransactionHistory.objects.filter(type_operation='Octroi de Crédit').aggregate(Sum('montant'))['montant__sum'] or 0
    paginator = Paginator(transactions, 15)
    page_number = request.GET.get('page', 1)
    try:
        transactions_paginees = paginator.page(page_number)
    except PageNotAnInteger:
        transactions_paginees = paginator.page(1)
    except EmptyPage:
        transactions_paginees = paginator.page(paginator.num_pages)

    context = {
        'transactions': transactions_paginees,
        'total_depots': total_depots,
        'total_retraits': total_retraits,
        'total_penalites': total_penalites,
        'total_credits': total_credits,
        'selected_type': type_op,
        'search_query': query,
    }
    return render(request, 'core/financial_report.html', context)


def get_brb_exchange_rates():
    """Récupère les taux USD et EUR depuis le site officiel de la BRB"""
    rates = {'USD': 'N/A', 'EUR': 'N/A'}
    try:
        url = "https://www.brb.bi/affichagetoustauxchange"
        headers = {'User-Agent': 'Mozilla/5.0'}
        response = requests.get(url, headers=headers, timeout=5)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            rows = soup.find_all('tr')
            for row in rows:
                text = row.get_text()
                if 'USD' in text:
                    cols = row.find_all('td')
                    if len(cols) >= 2:
                        rates['USD'] = cols[1].get_text(strip=True)
                if 'EUR' in text:
                    cols = row.find_all('td')
                    if len(cols) >= 2:
                        rates['EUR'] = cols[1].get_text(strip=True)
    except Exception as e:
        print("Erreur lors de la récupération des taux BRB:", e)
    return rates


def dashboard_view(request):
    if 'user_id' not in request.session and 'membre_id' not in request.session:
        return redirect('login')

    membre = None
    membre_id = request.session.get('membre_id')
    if membre_id:
        membre = get_object_or_404(Membres, id=membre_id)

    transactions = TransactionHistory.objects.values('type_operation').annotate(total=Sum('montant'))
    labels = [item['type_operation'] for item in transactions]
    data = [float(item['total'] or 0) for item in transactions]
    brb_rates = get_brb_exchange_rates()

    context = {
        'membre': membre,
        'user_type': request.session.get('user_type'),
        'chart_labels': json.dumps(labels),
        'chart_data': json.dumps(data),
        'brb_rates': brb_rates,
    }

    return render(request, 'core/dashboard.html', context)


def filtered_transactions_view(request):
    role = str(request.session.get('role', '')).lower()
    if 'membre_id' not in request.session or ('admin' not in role and 'gestionnaire' not in role):
        return redirect('login')
    transactions = TransactionHistory.objects.all().select_related('membre')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date:
        transactions = transactions.filter(date_transaction__date__gte=start_date)
    if end_date:
        transactions = transactions.filter(date_transaction__date__lte=end_date)
    total_montant = transactions.aggregate(Sum('montant'))['montant__sum'] or 0
    context = {
        'transactions': transactions, 'total_montant': total_montant, 'start_date': start_date or '', 'end_date': end_date or '',
    }
    return render(request, 'core/filtered_transactions.html', context)


def search_member_view(request):
    role = str(request.session.get('role', '')).lower()
    if 'membre_id' not in request.session or ('admin' not in role and 'gestionnaire' not in role):
        return redirect('login')
    query = request.GET.get('q', '')
    membres = []
    if query:
        filters = Q(nom__icontains=query) | Q(prenom__icontains=query) | Q(telephone__icontains=query)
        if query.isdigit():
            filters |= Q(id=int(query))
        membres = Membres.objects.filter(filters)
    return render(request, 'core/search_member.html', {'membres': membres, 'query': query})


def transaction_detail_view(request, transaction_id):
    transaction = get_object_or_404(
        TransactionHistory.objects.select_related('membre'), id=transaction_id
    )
    context = {
        'transaction': transaction,
    }
    return render(request, 'core/transaction_detail.html', context)


def export_transaction_pdf(request, transaction_id):
    transaction = get_object_or_404(TransactionHistory.objects.select_related('membre'), id=transaction_id)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="transaction_{transaction.id}.pdf"'
    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    p.setFont("Helvetica-Bold", 18)
    p.drawString(50, height - 50, "SACCO Connect - Reçu de Transaction")
    p.setFont("Helvetica", 10)
    p.drawString(50, height - 70, f"Date d'édition : {transaction.date_transaction.strftime('%d/%m/%Y %H:%M')}")
    p.setLineWidth(1)
    p.line(50, height - 85, width - 50, height - 85)
    y = height - 130
    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, y, f"Transaction N° : #{transaction.id}")
    y -= 30
    p.setFont("Helvetica", 11)
    p.drawString(50, y, f"Membre : {transaction.membre.nom} {transaction.membre.prenom}")
    y -= 25
    p.drawString(50, y, f"Téléphone : {transaction.membre.telephone}")
    y -= 25
    p.drawString(50, y, f"Type d'opération : {transaction.type_operation}")
    y -= 25
    p.drawString(50, y, f"Montant : {transaction.montant} BIF")
    y -= 25
    p.drawString(50, y, f"Statut : {transaction.statut}")
    y -= 35
    p.drawString(50, y, "Description :")
    y -= 20
    p.setFont("Helvetica-Oblique", 10)
    p.drawString(50, y, f"{transaction.description or 'Aucune description fournie.'}")
    p.setFont("Helvetica", 8)
    p.drawString(50, 50, "Document généré automatiquement par SACCO Connect.")
    p.showPage()
    p.save()
    return response


def edit_transaction_view(request, transaction_id):
    transaction = get_object_or_404(TransactionHistory, id=transaction_id)
    if request.method == 'POST':
        form = TransactionForm(request.POST, instance=transaction)
        if form.is_valid():
            form.save()
            messages.success(request, "La transaction a été mise à jour avec succès.")
            return redirect('transaction_detail', transaction_id=transaction.id)
    else:
        form = TransactionForm(instance=transaction)

    return render(request, 'core/edit_transaction.html', {
        'form': form, 'transaction': transaction
    })


@permission_required('core.delete_transactionhistory', raise_exception=True)
def delete_transaction_view(request, transaction_id):
    transaction = get_object_or_404(TransactionHistory, id=transaction_id)
    if request.method == 'POST':
        transaction.delete()
        messages.success(request, "La transaction a été supprimée avec succès.")
        return redirect('filtered_transactions')
    return render(request, 'core/confirm_delete.html', {'transaction': transaction})


@permission_required('auth.add_user', raise_exception=True)
def create_employee_view(request):
    if request.method == 'POST':
        form = EmployeeCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, f"L'employé {user.username} a été créé avec succès !")
            return redirect('dashboard')
    else:
        form = EmployeeCreationForm()
    return render(request, 'core/create_employee.html', {'form': form})


def list_loans_view(request):
    status_filter = request.GET.get('status')
    loans = Pret.objects.all().order_by('-date_demande')
    if status_filter:
        loans = loans.filter(statut=status_filter)
    context = {
        'loans': loans, 'current_status': status_filter
    }
    return render(request, 'core/loan_history.html', context)


def export_members_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="liste_membres.csv"'
    writer = csv.writer(response)
    writer.writerow(['ID', 'Nom', 'Email', 'Téléphone', 'Solde'])
    members = Membres.objects.all().values_list('id', 'nom', 'telephone', 'solde_epargne')
    for member in members:
        writer.writerow(member)
    return response


def request_loan_view(request):
    membre_id = request.session.get('membre_id')
    if not membre_id:
        return redirect('login')

    membre = get_object_or_404(Membres, id=membre_id)

    if request.method == 'POST':
        form = LoanRequestForm(request.POST, membre=membre)
        if form.is_valid():
            loan = form.save(commit=False)
            epargne_actuelle = membre.solde_epargne or 0
            plafond_emprunt = epargne_actuelle * 5

            if loan.montant > plafond_emprunt:
                messages.error(
                    request,
                    f"Montant non autorisé. Vous pouvez emprunter jusqu'à {plafond_emprunt:,} BIF (5x votre épargne de {epargne_actuelle:,} BIF)."
                )
                return render(request, 'core/request_loan.html', {'form': form, 'membre': membre})

            loan.membre = membre
            loan.statut = 'EN_ATTENTE'
            loan.taux_interet = 5.0
            loan.save()
            messages.success(request, "Votre demande de prêt a été soumise avec succès.")
            return redirect('dashboard')
    else:
        form = LoanRequestForm(membre=membre)

    return render(request, 'core/request_loan.html', {'form': form, 'membre': membre})


def get_filtered_loans(request):
    status_filter = request.GET.get('status')
    loans = Pret.objects.all().select_related('membre').order_by('-date_demande')
    if status_filter:
        loans = loans.filter(statut=status_filter)
    return loans


def export_loans_csv(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="historique_prets.csv"'
    response.write(u'\ufeff'.encode('utf8'))
    writer = csv.writer(response, delimiter=';')
    writer.writerow(['ID', 'Membre', 'Montant (BIF)', 'Taux (%)', 'Durée (mois)', 'Statut', 'Date'])
    loans = get_filtered_loans(request)
    for loan in loans:
        writer.writerow([
            loan.id, f"{loan.membre.nom} {loan.membre.prenom}", loan.montant, loan.taux_interet, loan.duree_mois,
            loan.get_statut_display(), loan.date_demande.strftime('%Y-%m-%d')
        ])
    return response


def export_loans_excel(request):
    if not openpyxl:
        return HttpResponse("La bibliothèque openpyxl n'est pas installée.", status=500)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rapport Prêts"
    headers = ['ID', 'Membre', 'Montant (BIF)', 'Taux (%)', 'Durée (mois)', 'Statut', 'Date']
    ws.append(headers)
    loans = get_filtered_loans(request)
    for loan in loans:
        ws.append([
            loan.id, f"{loan.membre.nom} {loan.membre.prenom}", float(loan.montant), float(loan.taux_interet),
            loan.duree_mois, loan.get_statut_display(), loan.date_demande.strftime('%Y-%m-%d')
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="historique_prets.xlsx"'
    wb.save(response)
    return response


def export_loans_pdf(request):
    if not pisa:
        return HttpResponse("La bibliothèque xhtml2pdf n'est pas installée.", status=500)

    loans = get_filtered_loans(request)
    status_filter = request.GET.get('status', 'Tous')
    context = {
        'loans': loans, 'current_status': status_filter,
    }
    html_string = render_to_string('core/pdf_loans_template.html', context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="historique_prets.pdf"'
    pisa_status = pisa.CreatePDF(html_string, dest=response)
    if pisa_status.err:
        return HttpResponse('Erreur lors de la génération du PDF', status=500)
    return response


def loan_detail_view(request, pret_id):
    membre_id = request.session.get('membre_id')
    role = str(request.session.get('role', '')).lower()
    if not membre_id and ('admin' not in role and 'gestionnaire' not in role and not request.user.is_authenticated):
        return redirect('login')
    if membre_id and 'admin' not in role and 'gestionnaire' not in role:
        pret = get_object_or_404(Pret, id=pret_id, membre_id=membre_id)
    else:
        pret = get_object_or_404(Pret, id=pret_id)
    amortissement = []
    if pret.statut == 'APPROUVE':
        amortissement = calculer_amortissement(pret.montant, pret.taux_interet, pret.duree_mois)
    return render(request, 'core/loan_detail.html', {'pret': pret, 'amortissement': amortissement})


@user_passes_test(is_agent_credit, login_url='/login/')
def valider_par_agent(request, pret_id):
    """Vue permettant à un agent de crédit de valider ou pré-valider une demande de prêt."""
    pret = get_object_or_404(Pret, id=pret_id)
    if request.method == 'POST':
        # Vous pouvez adapter le statut selon votre logique métier (ex: 'VALIDE_AGENT' ou 'EN_ATTENTE_MANAGER')
        action = request.POST.get('action')
        if action == 'valider':
            pret.statut = 'APPROUVE'  # ou un statut intermédiaire
            messages.success(request, f"Le prêt #{pret.id} a été validé par l'agent.")
        elif action == 'rejeter':
            pret.statut = 'REJETE'
            messages.warning(request, f"Le prêt #{pret.id} a été rejeté par l'agent.")
        pret.save()
        return redirect('agent_dashboard')

    return render(request, 'core/valider_pret_agent.html', {'pret': pret})


@user_passes_test(is_manager, login_url='/login/')
def approbation_finale_directeur(request, pret_id):
    """Vue permettant au manager / directeur d'accorder l'approbation finale d'un prêt."""
    pret = get_object_or_404(Pret, id=pret_id)
    if request.method == 'POST':
        action = request.POST.get('action', 'approuver')
        if action == 'approuver':
            pret.statut = 'APPROUVE'
            pret.date_approbation = timezone.now()
            messages.success(request,
                             f"Approbation finale accordée pour le prêt #{pret.id} ({pret.membre.nom} {pret.membre.prenom}).")
        elif action == 'rejeter':
            pret.statut = 'REJETE'
            pret.date_approbation = timezone.now()
            messages.warning(request, f"Le prêt #{pret.id} a été rejeté lors de l'approbation finale.")
        pret.save()
        return redirect('manager_dashboard')

    return render(request, 'core/approbation_directeur.html', {'pret': pret})


def security_pin_view(request):
    membre_id = request.session.get('membre_id')
    if not membre_id:
        return redirect('login')

    membre = get_object_or_404(Membres, id=membre_id)
    success_message = None
    error_message = None

    if request.method == 'POST' and 'update_pin' in request.POST:
        old_pin = request.POST.get('old_pin')
        new_pin = request.POST.get('new_pin')

        if str(membre.pin) == str(old_pin):
            membre.pin = new_pin
            membre.save()
            success_message = "Votre code PIN a été mis à jour avec succès."
        else:
            error_message = "L'ancien code PIN est incorrect."

    return render(request, 'core/security_pin.html', {
        'membre': membre,
        'success_message': success_message,
        'error_message': error_message
    })


def group_validate_loans_view(request):
    membre_id = request.session.get('membre_id')
    if not membre_id:
        return redirect('login')

    membre = get_object_or_404(Membres, id=membre_id)

    if membre.role != 'president':
        return redirect('dashboard')

    return render(request, 'core/group_validate_loans.html', {'membre': membre})


@login_required
def group_validate_loans(request):
    try:
        admin_membre = Membre.objects.get(user=request.user)
    except Membre.DoesNotExist:
        messages.error(request, "Profil membre introuvable.")

        return redirect('dashboard')

    if admin_membre.role not in ['president', 'secretaire']:
        messages.error(request, "Accès réservé aux présidents et secrétaires.")
        return redirect('dashboard')

    prets = Pret.objects.filter(membre__groupe=admin_membre.groupe, statut='EN_ATTENTE')
    context = {
        'prets': prets,
        'admin_membre': admin_membre,
    }
    return render(request, 'core/pending_loan.html', context)

@login_required
def update_loan_status(request, pret_id, action):
    admin_membre = get_object_or_404(Membre, user=request.user)

    if admin_membre.role not in ['president', 'secretaire']:
        messages.error(request, "Action non autorisée.")
        return redirect('dashboard')

    pret = get_object_or_404(Pret, id=pret_id, membre__groupe=admin_membre.groupe)

    if action == 'approuver':
        pret.statut = 'APPROUVE'
        pret.save()
        messages.success(request, f"Le prêt de #{pret.id} a été approuvé avec succès.")
    elif action == 'rejeter':
        pret.statut = 'REJETE'
        pret.save()
        messages.warning(request, f"Le prêt de #{pret.id} a été rejeté.")

    return redirect('group_validate_loans')


pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

def admin_reset_pin(request):
    if request.method == "POST":
        membre_id = request.POST.get("membre_id")
        nouveau_pin = request.POST.get("nouveau_pin")

        if not nouveau_pin or len(nouveau_pin) < 4:
            messages.error(request, "Le code PIN doit contenir au moins 4 caractères.")
            return redirect('manager_dashboard')

        hashed_pin = pwd_context.hash(nouveau_pin)

        try:
            with connection.cursor() as cursor:
                cursor.execute("UPDATE membres SET pin = %s WHERE id = %s", [hashed_pin, membre_id])

            messages.success(request, f"✅ Code PIN mis à jour avec succès pour le membre #{membre_id}.")
        except Exception as e:
            messages.error(request, f"Erreur lors de la réinitialisation : {e}")

    return redirect('manager_dashboard')


def admin_toggle_status(request):
    if request.method == "POST":
        target_id = request.POST.get("target_id")
        action_type = request.POST.get("action_type")

        new_status = False if action_type == "desactiver" else True

        try:
            with connection.cursor() as cursor:
                cursor.execute("UPDATE membres SET is_active = %s WHERE id = %s", [new_status, target_id])

            status_txt = "désactivé" if not new_status else "réactivé"
            messages.success(request, f"✅ Membre #{target_id} {status_txt} avec succès.")
        except Exception as e:
            messages.error(request, f"Erreur lors du changement de statut : {e}")

    return redirect('manager_dashboard')


@require_POST
def creer_groupe_view(request):
    nom = request.POST.get('nom_groupe')
    if nom:
        Groupe.objects.create(nom=nom)
    return redirect('manager_dashboard')


@require_POST
def archiver_groupe_view(request, groupe_id):
    groupe = get_object_or_404(Groupe, id=groupe_id)
    groupe.est_archive = True
    groupe.save()
    return redirect('manager_dashboard')


@require_POST
def restaurer_groupe_view(request, groupe_id):
    groupe = get_object_or_404(Groupe, id=groupe_id)
    groupe.est_archive = False
    groupe.save()
    return redirect('manager_dashboard')


def retirer_groupe_view(request, membre_id):
    if request.method == 'POST':
        membre = get_object_or_404(Membres, id=membre_id)
        membre.groupe = None  # Dissocie le groupe du membre
        membre.save()
    return redirect(request.META.get('HTTP_REFERER', 'manager_dashboard_view'))


@require_POST
def changer_groupe_membre_view(request, membre_id):
    membre = get_object_or_404(Membre, id=membre_id)
    nouveau_groupe_id = request.POST.get('nouveau_groupe_id')
    if nouveau_groupe_id:
        membre.groupe_id = nouveau_groupe_id
        membre.save()
    return redirect('manager_dashboard')


@require_POST
def retirer_groupe_view(request, membre_id):
    if request.method == 'POST':
        membre = get_object_or_404(Membres, id=membre_id)
        membre.groupe = None
        membre.save()
    return redirect(request.META.get('HTTP_REFERER', 'manager_dashboard_view'))


@require_POST
def creer_groupe_view(request):
    nom_saisi = request.POST.get('nom_groupe')
    if nom_saisi:
        Groupe.objects.create(nom_groupe=nom_saisi)
    return redirect('manager_dashboard')


@require_POST
def changer_groupe_view(request, membre_id):
    membre = get_object_or_404(Membres, id=membre_id)
    nouveau_groupe_id = request.POST.get('nouveau_groupe_id')

    if nouveau_groupe_id:
        nouveau_groupe = get_object_or_404(Groupe, id=nouveau_groupe_id)
        membre.groupe = nouveau_groupe
        membre.save()

    return redirect('manager_dashboard')


@require_POST
def assigner_partenaire_groupe_view(request, groupe_id):
    groupe = get_object_or_404(Groupes, id=groupe_id)
    partenaire_id = request.POST.get('partenaire_id')

    if partenaire_id:
        groupe.partenaire_id = partenaire_id if partenaire_id != "" else None
        groupe.save()

    return redirect('manager_dashboard')


def saisie_hebdomadaire_view(request):
    if request.method == 'POST':
        date_reunion = request.POST.get('date_reunion')
        groupe_id = request.POST.get('groupe_id')
        membre_ids = request.POST.getlist('membre_ids')
        heure_actuelle = datetime.now().strftime("%H:%M:%S")

        try:
            with transaction.atomic():
                for m_id in membre_ids:
                    presence = request.POST.get(f'presence_{m_id}', 'P')
                    epargne = float(request.POST.get(f'epargne_{m_id}', 0))
                    sociale = float(request.POST.get(f'sociale_{m_id}', 0))
                    amende = float(request.POST.get(f'amende_{m_id}', 0))
                    HistoriqueEpargne.objects.create(
                        membre_id=m_id, groupe_id=groupe_id if groupe_id else None, date_reunion=date_reunion,
                        epargne=epargne, caisse_sociale=sociale, amende=amende, status_presence=presence,
                        heure_enregistrement=heure_actuelle
                    )

                    membre = Membres.objects.get(id=m_id)
                    membre.solde_epargne = (membre.solde_epargne or 0) + epargne
                    membre.caisse_sociale = (membre.caisse_sociale or 0) + sociale
                    membre.status_presence = presence
                    membre.save()

            messages.success(request, f"✅ Séance du {date_reunion} enregistrée avec succès !")
        except Exception as e:
            messages.error(request, f"❌ Erreur lors de l'enregistrement : {e}")

    return redirect(request.META.get('HTTP_REFERER', 'manager_dashboard_view'))


def get_context_caisse_sociale(selected_gid):
    if not selected_gid:
        return {}

    membres = Membres.objects.filter(groupe_id=selected_gid, is_active=True).order_by('nom')
    historique_qs = HistoriqueEpargne.objects.filter(groupe_id=selected_gid).order_by('date_reunion')
    dates_reunions = list(historique_qs.values_list('date_reunion', flat=True).distinct())
    pivot_sociale = []
    totaux_par_date = [0] * len(dates_reunions)
    total_report_anterieur = 0
    grand_total_caisse_sociale = 0

    for m in membres:
        versements = HistoriqueEpargne.objects.filter(membre_id=m.id, groupe_id=selected_gid)
        dict_versements = {v.date_reunion: (v.caisse_sociale or 0) for v in versements}
        montants = []
        somme_session = 0
        for idx, date_r in enumerate(dates_reunions):
            m_val = dict_versements.get(date_r, 0)
            montants.append(m_val)
            somme_session += m_val
            totaux_par_date[idx] += m_val

        report_anterieur = (m.caisse_sociale or 0) - somme_session
        if report_anterieur < 0:
            report_anterieur = 0

        total_cumul = report_anterieur + somme_session
        total_report_anterieur += report_anterieur
        grand_total_caisse_sociale += total_cumul

        pivot_sociale.append({
            'membre_id': m.id, 'nom_complet': f"{m.prenom or ''} {m.nom}".strip(), 'report_anterieur': report_anterieur,
            'montants': montants, 'total_cumul': total_cumul,
        })

    return {
        'dates_reunions': [d.strftime('%d/%m/%Y') if hasattr(d, 'strftime') else d for d in dates_reunions],
        'pivot_sociale': pivot_sociale, 'totaux_par_date': totaux_par_date, 'total_report_anterieur': total_report_anterieur,
        'grand_total_caisse_sociale': grand_total_caisse_sociale, 'total_cotisations_mois': sum(totaux_par_date),
        'total_emprunts_socials': 0
    }


def tout_recuperer_donnees_caisse(request):
    """
    Fonction utilitaire pour récupérer et calculer les données de la caisse sociale
    en fonction des filtres (groupe, membre, mois) passés dans la requête.
    """
    selected_gid = request.GET.get('selected_gid', '')
    selected_member_id = request.GET.get('selected_member_id', '')
    mois_filtre = request.GET.get('mois_filtre', datetime.now().strftime('%Y-%m'))
    annee, mois = map(int, mois_filtre.split('-'))
    prefixe_mois = f"{annee}-{mois:02d}"
    membres_qs = Membres.objects.filter(is_active=True)
    if selected_gid:
        membres_qs = membres_qs.filter(groupe_id=selected_gid)
    if selected_member_id:
        membres_qs = membres_qs.filter(id=selected_member_id)

    dates_reunions_brutes = HistoriqueEpargne.objects.filter(
        date_reunion__startswith=prefixe_mois
    ).values_list('date_reunion', flat=True).distinct().order_by('date_reunion')
    dates_reunions_formatees = [f"{d[-2:]}/{d[5:7]}" for d in dates_reunions_brutes if d]
    pivot_sociale = []

    for m in membres_qs:
        report_anterieur = HistoriqueEpargne.objects.filter(
            membre_id=m.id,
            date_reunion__lt=f"{prefixe_mois}-01"
        ).aggregate(Sum('caisse_sociale'))['caisse_sociale__sum'] or 0
        montants_mois = []
        sum_mois = 0
        for d_brute in dates_reunions_brutes:
            valeur = HistoriqueEpargne.objects.filter(
                membre_id=m.id,
                date_reunion=d_brute
            ).aggregate(Sum('caisse_sociale'))['caisse_sociale__sum'] or 0
            montants_mois.append(valeur)
            sum_mois += valeur

        cumul_actuel = report_anterieur + sum_mois

        pivot_sociale.append({
            'id': m.id, 'nom_complet': f"{m.prenom or ''} {m.nom}".strip(), 'report_anterieur': report_anterieur,
            'montants': montants_mois, 'total_mois': sum_mois, 'cumul_actuel': cumul_actuel,
        })

    return pivot_sociale, dates_reunions_formatees, mois_filtre


def export_caisse_sociale_excel(request):
    """Génère un fichier Excel stylisé du registre Caisse Sociale."""
    pivot_sociale, dates_reunions, mois_filtre = tout_recuperer_donnees_caisse(request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Caisse Sociale {mois_filtre}"
    headers = ["ID", "Membre", "Report Antérieur (BIF)"] + [f"📅 {d}" for d in dates_reunions] + ["Total Mois (BIF)", "Cumul Actuel (BIF)"]
    ws.append(headers)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for line in pivot_sociale:
        row = [
            line['id'], line['nom_complet'], line['report_anterieur']
        ] + line['montants'] + [line['total_mois'], line['cumul_actuel']]
        ws.append(row)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Caisse_Sociale_{mois_filtre}.xlsx"'
    wb.save(response)
    return response


def export_caisse_sociale_pdf(request):
    """Génère un document PDF imprimable du registre."""
    pivot_sociale, dates_reunions, mois_filtre = tout_recuperer_donnees_caisse(request)
    context = {
        'pivot_sociale': pivot_sociale, 'dates_reunions': dates_reunions, 'mois_filtre': mois_filtre,
    }

    html_string = render_to_string('pdf/caisse_sociale_pdf.html', context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Caisse_Sociale_{mois_filtre}.pdf"'
    pisa_status = pisa.CreatePDF(
        html_string, dest=response, link_callback=link_callback
    )

    if pisa_status.err:
        return HttpResponse('Erreur lors de la génération du PDF', status=500)
    return response


def export_members_excel(request):
    search_query = request.GET.get('q', '')
    membres_qs = Membres.objects.all()
    if search_query:
        membres_qs = membres_qs.filter(
            Q(nom__icontains=search_query) |
            Q(prenom__icontains=search_query) |
            Q(telephone__icontains=search_query)
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liste des Membres"
    headers = ["ID", "Nom & Prénom", "Sexe", "Âge", "Téléphone", "CNI", "Adresse Complète", "Solde Épargne (BIF)",
               "Crédit en cours (BIF)"]
    ws.append(headers)
    header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for m in membres_qs:
        colline = getattr(m, 'colline', None) or '-'
        quartier = getattr(m, 'quartier', None) or '-'
        avenue = getattr(m, 'avenue', None) or '-'
        numero_maison = getattr(m, 'numero_maison', None) or 'N/A'
        adresse = f"{colline}, {quartier} - Av. {avenue} (N° {numero_maison})"

        row = [
            m.id, f"{getattr(m, 'nom', '')} {getattr(m, 'prenom', '')}".strip(), getattr(m, 'sexe', None) or "N/A",
            getattr(m, 'age', None) or "N/A", getattr(m, 'telephone', ''), getattr(m, 'cni', None) or "N/A",
            adresse, getattr(m, 'solde_epargne', 0) or 0, getattr(m, 'credit_en_cours', 0) or 0
        ]
        ws.append(row)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Membres_Sacco.xlsx"'
    wb.save(response)

    return response


def link_callback(uri, rel):
    """
    Convertit les URLs web statiques en chemins de fichiers absolus locaux
    pour que xhtml2pdf puisse intégrer les images.
    """
    result = uri
    if uri.startswith(settings.STATIC_URL):
        path = finders.find(uri.replace(settings.STATIC_URL, ""))
        if path:
            if isinstance(path, (list, tuple)):
                path = path[0]
            result = path
    elif uri.startswith(settings.MEDIA_URL):
        result = os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ""))

    return result


def export_transactions_pdf(request):
    """Génère un rapport PDF des transactions avec filtres actifs."""
    search_query = request.GET.get('q', '')
    selected_type = request.GET.get('type_op', '')
    transactions = TransactionHistory.objects.all().order_by('-date_transaction')

    if search_query:
        transactions = transactions.filter(
            models.Q(membre__nom__icontains=search_query) |
            models.Q(membre__prenom__icontains=search_query)
        )
    if selected_type:
        transactions = transactions.filter(type_operation=selected_type)

    context = {
        'transactions': transactions,
    }

    html_string = render_to_string('pdf/transactions_pdf.html', context, request=request)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="Rapport_Financier_Transactions.pdf"'

    pisa_status = pisa.CreatePDF(
        html_string,
        dest=response,
        link_callback=link_callback
    )

    if pisa_status.err:
        return HttpResponse('Erreur lors de la génération du PDF', status=500)
    return response


def admin_planifier_reunion_view(request):
    if request.method == 'POST':
        groupe_id = request.POST.get('groupe_id')
        prochaine_reunion = request.POST.get('prochaine_reunion')
        annonce_message = request.POST.get('annonce_message')

        # Logique d'enregistrement (ex: sauvegarder dans un modèle Groupe ou une table d'annonces)
        # Exemple:
        # groupe = Groupes.objects.get(id=groupe_id)
        # groupe.prochaine_reunion = prochaine_reunion
        # groupe.annonce_message = annonce_message
        # groupe.save()

        messages.success(request, "La date de la prochaine réunion a été planifiée avec succès pour le groupe.")
    return redirect('manager_dashboard')
